"""Import additif (incremental_import.py / push_scored_import.push_incremental_import).

Régression pour l'incident réel : un import Éducation physique seule pour des
élèves déjà importés cette année (autre matière) échouait sur
students_pseudo_academic_year_unique, et un import "premier passage" à une
seule matière ne produisait aucun signal de risque exploitable. Ces tests
n'ouvrent jamais de connexion réseau réelle : `FakeSupabaseClient` simule les
tables students/grades/recommendations en mémoire, en ne supportant que les
filtres PostgREST réellement utilisés par le code (`eq.`, `in.(...)`)."""
import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import incremental_import
from anonymization.anonymize import get_or_create_salt, pseudonymize_code
from persistence.push_scored_import import push_incremental_import

MODELS_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "artifacts", "models")


class FakeSupabaseClient:
    """Double de test pour SupabaseRestClient : mêmes signatures publiques
    (select/insert/delete), stockage en mémoire, pas de réseau."""

    def __init__(self, seed: dict[str, list[dict]] | None = None):
        self.tables: dict[str, list[dict]] = {k: list(v) for k, v in (seed or {}).items()}
        self.calls: list[tuple[str, str]] = []  # (method, table), dans l'ordre

    @staticmethod
    def _matches(row: dict, key: str, expr: str) -> bool:
        if expr.startswith("eq."):
            return str(row.get(key)) == expr[3:]
        if expr.startswith("in.(") and expr.endswith(")"):
            values = expr[4:-1].split(",")
            return str(row.get(key)) in values
        raise NotImplementedError(f"Filtre non supporté par FakeSupabaseClient : {key}={expr}")

    def select(self, table: str, params: dict[str, str]) -> list[dict]:
        self.calls.append(("select", table))
        rows = self.tables.get(table, [])
        for key, expr in params.items():
            if key == "select" or key.startswith("order") or key == "limit":
                continue
            rows = [r for r in rows if self._matches(r, key, expr)]
        return rows

    def insert(self, table: str, rows: list[dict], chunk_size: int = 500, upsert_on_conflict: str | None = None) -> int:
        self.calls.append(("insert", table))
        if not rows:
            return 0
        existing = self.tables.setdefault(table, [])
        if upsert_on_conflict:
            conflict_cols = upsert_on_conflict.split(",")
            new_keys = {tuple(r[c] for c in conflict_cols) for r in rows}
            existing[:] = [r for r in existing if tuple(r[c] for c in conflict_cols) not in new_keys]
        existing.extend(rows)
        return len(rows)

    def delete(self, table: str, filters: dict[str, str]) -> None:
        self.calls.append(("delete", table))
        rows = self.tables.get(table, [])
        keep = []
        for r in rows:
            matched = all(self._matches(r, key, expr) for key, expr in filters.items())
            if not matched:
                keep.append(r)
        self.tables[table] = keep


def _build_massar_file(path: str, matiere_ar: str, student_code: str, nom: str, notes: dict) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="القسم")
    ws.cell(row=1, column=2, value="1APIC-1")
    ws.cell(row=2, column=1, value="المادة")
    ws.cell(row=2, column=2, value=matiere_ar)
    ws.cell(row=3, column=2, value="ID")
    ws.cell(row=3, column=3, value="Code")
    ws.cell(row=3, column=4, value="Nom")
    ws.cell(row=3, column=7, value="الفرض الأول")
    ws.cell(row=3, column=8, value="الفرض الثاني")
    ws.cell(row=5, column=2, value="1")
    ws.cell(row=5, column=3, value=student_code)
    ws.cell(row=5, column=4, value=nom)
    ws.cell(row=5, column=7, value=notes["c1"])
    ws.cell(row=5, column=8, value=notes["c2"])
    wb.save(path)


@pytest.fixture(autouse=True)
def _fake_client(monkeypatch):
    """Remplace SupabaseRestClient par un double en mémoire pour tout le
    module — aucun test ici ne doit toucher le vrai Supabase."""
    holder: dict[str, FakeSupabaseClient] = {}

    def factory(*args, **kwargs):
        client = holder.setdefault("client", FakeSupabaseClient())
        return client

    monkeypatch.setattr(incremental_import, "SupabaseRestClient", factory)
    yield holder
    holder.clear()


def _pseudo(student_code: str) -> str:
    return pseudonymize_code(student_code, get_or_create_salt())


def test_first_pass_all_students_new(local_tmp_path, _fake_client):
    raw_dir = local_tmp_path / "raw"
    raw_dir.mkdir()
    _build_massar_file(
        str(raw_dir / "Export_28847E_1APIC-1_MATHEMATIQUES_20260101000000.xlsx"),
        "الرياضيات", "TESTCODE001", "Élève Test", {"c1": 12.0, "c2": 14.0},
    )
    result = incremental_import.run_incremental_import(str(raw_dir), "2099/2100", models_dir=MODELS_DIR)

    assert result["n_students"] == 1
    assert result["n_students_nouveaux"] == 1
    assert result["n_students_completes"] == 0
    assert result["profile"].iloc[0]["nb_matieres_suivies"] == 1


def test_second_subject_merges_with_existing_student(local_tmp_path, _fake_client):
    student_code = "TESTCODE002"
    pseudo = _pseudo(student_code)
    existing_id = "11111111-1111-1111-1111-111111111111"

    # Seed : cet élève a déjà une ligne `students` + une matière en `grades`
    # pour 2099/2100 (import précédent, Mathématiques : moyenne (12+14)/2=13).
    client = FakeSupabaseClient(
        seed={
            "students": [
                {
                    "id": existing_id, "student_pseudo": pseudo, "academic_year": "2099/2100",
                    "niveau": "1APIC", "classe": "1APIC-1", "age": 12,
                }
            ],
            "grades": [
                {
                    "student_id": existing_id, "subject_code": "MATHEMATIQUES",
                    "c1": 12.0, "c2": 14.0, "c3": None, "c4": None, "activites": None,
                    "c1_colonne_existe": True, "c2_colonne_existe": True, "c3_colonne_existe": False,
                    "c4_colonne_existe": False, "activites_colonne_existe": False,
                    "moyenne_matiere": 13.0, "tendance_matiere": None, "remarque_fr": None,
                }
            ],
        }
    )
    _fake_client["client"] = client

    raw_dir = local_tmp_path / "raw"
    raw_dir.mkdir()
    _build_massar_file(
        str(raw_dir / "Export_28847E_1APIC-1_PHYSIQUE CHIMIE_20260101000000.xlsx"),
        "الفيزياء والكيمياء", student_code, "Élève Test", {"c1": 8.0, "c2": 10.0},
    )

    result = incremental_import.run_incremental_import(str(raw_dir), "2099/2100", models_dir=MODELS_DIR)

    assert result["n_students"] == 1
    assert result["n_students_nouveaux"] == 0
    assert result["n_students_completes"] == 1
    assert result["pseudo_to_id"][pseudo] == existing_id  # id existant réutilisé, pas régénéré

    row = result["profile"].iloc[0]
    assert row["nb_matieres_suivies"] == 2
    # moyenne_generale = moyenne des deux moyennes de matière : (13.0 + 9.0) / 2 = 11.0
    assert row["moyenne_generale"] == pytest.approx(11.0)


def test_push_incremental_import_clears_old_recommendations_before_reinsert(local_tmp_path, _fake_client, monkeypatch):
    existing_id = "22222222-2222-2222-2222-222222222222"
    pseudo = "abcdef0123456789"
    client = FakeSupabaseClient(
        seed={
            "recommendations": [
                {"id": "old-rec", "student_id": existing_id, "dataset_id": "old-ds", "priorite": 1,
                 "type": "ancien", "justification": "x", "action": "y", "matieres_concernees": [],
                 "profil": None, "tendance_previsionnelle_moyenne_predite": None}
            ]
        }
    )

    import pandas as pd

    profile = pd.DataFrame(
        [{
            "student_pseudo": pseudo, "niveau": "1APIC", "classe": "1APIC-1",
            "nb_matieres_suivies": 2, "moyenne_generale": 11.0,
            "dispersion_intermatiere": 1.5, "tendance_globale": None,
            "remarque_encodee": None,
            "cluster_id": None, "cluster_label": None, "pca_1": None, "pca_2": None,
            "a_risque": False, "a_risque_predit": False, "probabilite_risque": 0.1,
            "moyenne_generale_predite": 11.0,
        }]
    )
    identity = pd.DataFrame([{"student_pseudo": pseudo, "nom_complet": "Élève Test", "age": 12}])
    result = {
        "academic_year": "2099/2100",
        "profile": profile,
        "identity_mapping": identity,
        "pseudo_to_id": {pseudo: existing_id},
        "notes_long_new": pd.DataFrame(columns=[
            "student_pseudo", "matiere", "c1", "c2", "c3", "c4", "activites",
            "c1_colonne_existe", "c2_colonne_existe", "c3_colonne_existe",
            "c4_colonne_existe", "activites_colonne_existe",
            "moyenne_matiere", "n_composantes", "tendance_matiere", "remarque",
        ]),
        "notes_long_aggregated": pd.DataFrame(columns=["student_pseudo", "matiere"]),
        "recommendations": [
            {"student_pseudo": pseudo, "priorite": 2, "type": "nouveau",
             "justification": "j", "action": "a", "matieres_concernees": []}
        ],
        "n_files_discovered": 0, "n_files_parsed_ok": 0, "n_files_quarantined": 0,
        "n_students": 1, "n_students_nouveaux": 0, "n_students_completes": 1,
        "n_anomalies_bornes": 0, "n_doublons": 0, "coverage_counts": {},
        "niveaux": ["1APIC"], "classes": ["1APIC-1"], "matieres": [],
    }

    import persistence.push_scored_import as psi

    def factory(*a, **k):
        return client

    monkeypatch.setattr(psi, "SupabaseRestClient", factory)
    push_incremental_import(result, label="Test")

    delete_idx = client.calls.index(("delete", "recommendations"))
    insert_idx = max(i for i, c in enumerate(client.calls) if c == ("insert", "recommendations"))
    assert delete_idx < insert_idx
    remaining = client.tables["recommendations"]
    assert all(r.get("type") != "ancien" for r in remaining)
    assert any(r.get("type") == "nouveau" for r in remaining)
