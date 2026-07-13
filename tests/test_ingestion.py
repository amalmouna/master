"""Étape A — parsing : matières hors périmètre modélisé.

Régression pour le bug réel observé en production : un export Massar pour une
matière absente de MATIERE_AR_TO_CODE (ex. Informatique, "المعلوميات", avant
l'ajout de son entrée) laissait passer le code arabe brut comme
`matiere`/`subject_code` jusqu'à la persistance, où il violait la contrainte
de clé étrangère `grades_subject_code_fkey` (aucune ligne correspondante dans
`subjects`). Le fichier doit être mis en quarantaine au parsing, comme un
fichier illisible. On utilise ici un libellé arabe fictif pour tester ce cas
— "المعلوميات" est désormais reconnu (mappé vers INFORMATIQUE, présent dans
`subjects`) et ne doit donc plus déclencher ce chemin."""
import os
import shutil
import tempfile

import openpyxl
import pytest

from ingestion.parse_massar import parse_file
from score_import import BLOCKING_ISSUE_CODES


@pytest.fixture
def tmp_path():
    d = tempfile.mkdtemp(prefix="test_ingestion_")
    try:
        yield d
    finally:
        shutil.rmtree(d, ignore_errors=True)


def _build_massar_like_file(path: str, matiere_ar: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="المادة")
    ws.cell(row=1, column=2, value=matiere_ar)
    ws.cell(row=2, column=2, value="ID")
    ws.cell(row=2, column=3, value="Code")
    ws.cell(row=2, column=4, value="Nom")
    ws.cell(row=2, column=7, value="الفرض الأول")
    ws.cell(row=4, column=2, value="1")
    ws.cell(row=4, column=3, value="C001")
    ws.cell(row=4, column=4, value="Élève Un")
    ws.cell(row=4, column=7, value=12.0)
    wb.save(path)


def test_matiere_connue_ne_declenche_aucun_probleme(tmp_path):
    path = os.path.join(tmp_path, "Export_28847E_1APIC-1_LANGUE ARABE_20260101000000.xlsx")
    _build_massar_like_file(path, "اللغة العربية")
    result = parse_file(path)
    assert result.matiere_content == "LANGUE ARABE"
    assert not any(i["code"] == "MATIERE_HORS_PERIMETRE" for i in result.issues)


def test_informatique_est_mappee_vers_son_code_canonique(tmp_path):
    # Régression exacte du bug rapporté : "المعلوميات" doit résoudre vers le
    # code déjà présent dans `subjects` (INFORMATIQUE), jamais rester en arabe brut.
    path = os.path.join(tmp_path, "Export_28847E_2APIC-1_INFORMATIQUE_20260101000000.xlsx")
    _build_massar_like_file(path, "المعلوميات")
    result = parse_file(path)
    assert result.matiere_content == "INFORMATIQUE"
    assert not any(i["code"] == "MATIERE_HORS_PERIMETRE" for i in result.issues)


@pytest.mark.parametrize(
    "matiere_ar,code_attendu",
    [
        ("التربية الإسلامية", "EDUCATION ISLAMIQUE"),
        ("التكنولوجيا", "TECHNOLOGIE"),
        ("التربية البدنية", "EDUCATION PHYSIQUE"),
    ],
)
def test_matieres_ajoutees_sans_fichier_reel_sont_mappees(tmp_path, matiere_ar, code_attendu):
    # Libellés non vérifiés contre un vrai export Massar (aucun des 75 fichiers
    # de data/raw/ ne couvre ces matières) — ce test fige seulement le mapping
    # ajouté, pas une garantie que Massar utilise exactement ce texte.
    path = os.path.join(tmp_path, f"Export_28847E_2APIC-1_{code_attendu}_20260101000000.xlsx")
    _build_massar_like_file(path, matiere_ar)
    result = parse_file(path)
    assert result.matiere_content == code_attendu
    assert not any(i["code"] == "MATIERE_HORS_PERIMETRE" for i in result.issues)


def test_matiere_inconnue_est_signalee_et_bloquante(tmp_path):
    matiere_inconnue = "مادة اختبار غير معروفة"  # libellé fictif, jamais dans MATIERE_AR_TO_CODE
    path = os.path.join(tmp_path, "Export_28847E_2APIC-1_INCONNUE_20260101000000.xlsx")
    _build_massar_like_file(path, matiere_inconnue)
    result = parse_file(path)

    assert result.matiere_content == matiere_inconnue  # code brut, jamais un code canonique inventé
    matiere_issues = [i for i in result.issues if i["code"] == "MATIERE_HORS_PERIMETRE"]
    assert len(matiere_issues) == 1
    assert matiere_issues[0]["detail"] == matiere_inconnue

    # C'est bien ce code qui fait quarantiner le fichier en amont de la persistance
    # (score_import._ingest_and_clean / pipeline_run.run) — sinon ce texte brut finirait
    # comme subject_code et violerait la FK grades_subject_code_fkey en base.
    assert "MATIERE_HORS_PERIMETRE" in BLOCKING_ISSUE_CODES
