"""Étape A — parsing robuste d'un export Massar (.xlsx) en enregistrements élève x matière.

Un fichier = une classe x une matière x un semestre. La disposition des colonnes de
notes varie d'un fichier à l'autre (4 variantes constatées sur les 75 fichiers réels :
présence ou non de C4 et/ou Activités) : les colonnes sont donc résolues par libellé
arabe dans la ligne d'en-tête, jamais par position fixe.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field

import openpyxl

# Libellés des composantes de note, dans l'ordre où elles apparaissent dans les fichiers.
COMPONENT_LABELS = [
    ("الفرض الأول", "c1"),
    ("الفرض الثاني", "c2"),
    ("الفرض الثالث", "c3"),
    ("الفرض الرابع", "c4"),
    ("الأنشطة المندمجة", "activites"),
]
REMARK_LABEL = "ملاحظات الأستاذ"
ID_LABEL = "ID"

# Matière (texte Massar en arabe) -> code canonique utilisé dans tout le projet.
MATIERE_AR_TO_CODE = {
    "الرياضيات": "MATHEMATIQUES",
    "الفيزياء والكيمياء": "PHYSIQUE CHIMIE",
    "علوم الحياة والأرض": "SC. DE LA VIE ET DE LA TERRE",
    "اللغة العربية": "LANGUE ARABE",
    "اللغة الفرنسية": "LANGUE FRANCAISE",
    "اللغة الإنجليزية": "LANGUE ANGLAISE",
    "الاجتماعيات": "HISTOIRE GEOGRAPHIE",
}

# Métadonnées d'en-tête : libellé arabe -> nom de champ. La valeur est cherchée en
# balayant vers la droite depuis la cellule du libellé (l'écart varie selon les
# fichiers, d'où un balayage plutôt qu'un décalage fixe).
METADATA_LABELS = {
    "القسم": "classe_content",
    "المادة": "matiere_content",
    "المستوى": "niveau_content",
    "الاستاذ": "enseignant",
    "الدورة": "session",
    "السنة الدراسية": "annee_scolaire",
    "مؤسسة": "etablissement",
}

FILENAME_RE = re.compile(
    r"Export_(?P<etab>[A-Za-z0-9]+)_(?P<classe>[0-9][A-Za-z]+-[0-9]+)_(?P<matiere>.+?)_(?P<horodatage>\d+)\.xlsx$",
    re.IGNORECASE,
)


@dataclass
class ParsedFile:
    source_file: str
    classe_filename: str | None = None
    matiere_filename: str | None = None
    classe_content: str | None = None
    matiere_content: str | None = None
    niveau: str | None = None
    metadata: dict = field(default_factory=dict)
    component_columns: dict = field(default_factory=dict)
    records: list = field(default_factory=list)
    issues: list = field(default_factory=list)
    ok: bool = True

    def add_issue(self, code: str, detail: str):
        self.issues.append({"code": code, "detail": detail})


def parse_filename(path: str) -> dict:
    name = os.path.basename(path)
    m = FILENAME_RE.match(name)
    if not m:
        return {}
    return {
        "etablissement": m.group("etab"),
        "classe": m.group("classe"),
        "matiere": m.group("matiere").strip(),
        "horodatage": m.group("horodatage"),
    }


def _find_header_row(ws, max_scan_rows: int = 40) -> int | None:
    for r in range(1, max_scan_rows + 1):
        v = ws.cell(row=r, column=2).value
        if v is not None and str(v).strip() == ID_LABEL:
            return r
    return None


def _find_metadata_value(ws, header_row: int, label: str, max_lookahead: int = 5):
    """Cherche `label` dans la zone métadonnées (au-dessus de l'en-tête) et renvoie
    la première cellule non vide trouvée en balayant vers la droite sur la même ligne."""
    for r in range(1, header_row):
        for c in range(1, 18):
            v = ws.cell(row=r, column=c).value
            if v is not None and label in str(v):
                for c2 in range(c + 1, min(c + 1 + max_lookahead, 19)):
                    v2 = ws.cell(row=r, column=c2).value
                    if v2 is not None and str(v2).strip():
                        return str(v2).strip()
    return None


def _resolve_component_columns(ws, header_row: int) -> dict:
    """Associe chaque libellé de composante à l'indice de colonne (1-based) qui
    contient la note (première occurrence du libellé, la seconde étant la colonne
    d'absence, toujours vide dans ce jeu de données)."""
    col_map = {}
    seen_labels = set()
    for c in range(1, 30):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        text = str(v).strip()
        for label, field_name in COMPONENT_LABELS:
            if text == label and label not in seen_labels:
                col_map[field_name] = c
                seen_labels.add(label)
        if text == REMARK_LABEL and REMARK_LABEL not in seen_labels:
            col_map["remarque"] = c
            seen_labels.add(REMARK_LABEL)
    return col_map


def parse_file(path: str) -> ParsedFile:
    result = ParsedFile(source_file=path)

    fn_meta = parse_filename(path)
    if not fn_meta:
        result.add_issue("FILENAME_UNPARSEABLE", os.path.basename(path))
    else:
        result.classe_filename = fn_meta["classe"]
        result.matiere_filename = fn_meta["matiere"]

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
    except Exception as exc:  # fichier corrompu / illisible
        result.ok = False
        result.add_issue("FILE_UNREADABLE", str(exc))
        return result

    header_row = _find_header_row(ws)
    if header_row is None:
        result.ok = False
        result.add_issue("HEADER_NOT_FOUND", "Cellule 'ID' introuvable en colonne B")
        return result

    for label, field_name in METADATA_LABELS.items():
        val = _find_metadata_value(ws, header_row, label)
        result.metadata[field_name] = val
    result.classe_content = result.metadata.get("classe_content")
    matiere_ar = result.metadata.get("matiere_content")
    result.matiere_content = MATIERE_AR_TO_CODE.get(matiere_ar, matiere_ar)
    if matiere_ar is not None and matiere_ar not in MATIERE_AR_TO_CODE:
        # Matière hors du périmètre modélisé (les 7 matières de MATIERE_AR_TO_CODE) :
        # ex. Informatique. Distinct d'une matière au programme mais non suivie par
        # un élève donné (ça, c'est "non_au_programme" dans build_coverage_matrix) —
        # ici le fichier entier ne correspond à aucune matière connue du projet, donc
        # on quarantaine plutôt que de laisser passer un code brut arabe qui violerait
        # la contrainte de clé étrangère grades.subject_code à la persistance.
        result.add_issue("MATIERE_HORS_PERIMETRE", matiere_ar)

    # Recoupement nom de fichier <-> contenu.
    if result.classe_filename and result.classe_content and result.classe_filename != result.classe_content:
        result.add_issue(
            "CLASSE_MISMATCH",
            f"filename={result.classe_filename} content={result.classe_content}",
        )
    if (
        result.matiere_filename
        and result.matiere_content
        and result.matiere_filename.upper() != str(result.matiere_content).upper()
    ):
        result.add_issue(
            "MATIERE_MISMATCH",
            f"filename={result.matiere_filename} content={result.matiere_content}",
        )

    classe = result.classe_content or result.classe_filename
    matiere = result.matiere_content or result.matiere_filename
    if classe:
        m = re.match(r"([0-9][A-Za-z]+)-\d+", classe)
        result.niveau = m.group(1) if m else None

    col_map = _resolve_component_columns(ws, header_row)
    result.component_columns = col_map
    if not col_map:
        result.add_issue("NO_COMPONENT_COLUMNS", "Aucune composante de note détectée")

    data_start = header_row + 2  # ligne suivant la sous-ligne NOTE/ABSENCE
    r = data_start
    max_rows = ws.max_row
    while r <= max_rows:
        student_id_interne = ws.cell(row=r, column=2).value
        if student_id_interne is None or str(student_id_interne).strip() == "":
            break
        record = {
            "id_interne": student_id_interne,
            "student_code": ws.cell(row=r, column=3).value,
            "nom_complet": ws.cell(row=r, column=4).value,
            "dob_raw": ws.cell(row=r, column=6).value,
            "niveau": result.niveau,
            "classe": classe,
            "matiere": matiere,
            "c1": col_map.get("c1") and ws.cell(row=r, column=col_map["c1"]).value,
            "c2": col_map.get("c2") and ws.cell(row=r, column=col_map["c2"]).value,
            "c3": col_map.get("c3") and ws.cell(row=r, column=col_map["c3"]).value,
            "c4": col_map.get("c4") and ws.cell(row=r, column=col_map["c4"]).value,
            "activites": col_map.get("activites") and ws.cell(row=r, column=col_map["activites"]).value,
            "remarque": col_map.get("remarque") and ws.cell(row=r, column=col_map["remarque"]).value,
            "source_file": os.path.basename(path),
        }
        for comp in ("c1", "c2", "c3", "c4", "activites"):
            record[f"{comp}_colonne_existe"] = comp in col_map
        result.records.append(record)
        r += 1

    if not result.records:
        result.add_issue("NO_STUDENT_ROWS", f"Aucun élève lu à partir de la ligne {data_start}")

    return result
